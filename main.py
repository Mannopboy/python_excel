from openpyxl import load_workbook, Workbook

info = [{'name': 'Id', 'number': 'A1'}, {'name': 'Name', 'number': 'B1'}, {'name': 'Surname', 'number': 'C1'}]
data = [
    {
        'name': 'Mannopboy',
        'surname': 'Mukinboyev'
    },
    {
        'name': 'Elyor',
        'surname': 'Xamidulayev'
    },
    {
        'name': 'Xudoyor',
        'surname': 'Tursunov'
    },
    {
        'name': 'Shohruh',
        'surname': 'Shermetov'
    },
]

# workbook = load_workbook(filename="document.xlsx")
new_workbook = Workbook()

sheet = new_workbook.active

for item in info:
    sheet[item['number']] = item['name']
number = 1
for item in data:
    sheet[f'A{number + 1}'] = number
    number += 1
number = 2
for item in data:
    sheet[f'B{number}'] = item['name']
    number += 1
number = 2
for item in data:
    sheet[f'C{number}'] = item['surname']
    number += 1

new_workbook.save(filename="csv/output.xlsx")
